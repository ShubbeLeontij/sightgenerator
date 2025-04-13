#!/usr/bin/env python3
import shutil

import generator
import json
import os
import openpyxl
import argparse

__author__ = "Shubbe Leontij"
__version__ = "4.3"


def reader(MODE, sheets=None, _print=print, _input=input):
    """
    Function that reads the table and creates sights for every row. This is the core of whole program.
    :param MODE: output mode. Development - 0 ; Normal - 1 (default) ; Silent - 2 ; Full silent - 3
    :param sheets: list of sheet names that will be used. By default, all sheets will be used
    :param _print: output function
    :param _input: input function
    """
    def _output(string, severity=1):
        """
        Function that check should some text be moved to output or not.
        :param string: text that potentially should be moved to output
        :param severity: integer severity of this text
        """
        if MODE <= severity:
            _print(string)

    # Loading path from json and making it correct
    with open("settings.json", 'r') as f:
        wt_path = json.load(f)["path"].replace('\\', '/')
    if wt_path == "":
        wt_path = os.path.dirname(os.path.realpath("settings.json")) + "/UserSights/"
    else:
        wt_path_list = wt_path.split('/')
        wt_path = '/' if wt_path[0] == '/' else ""
        for folder in wt_path_list:
            if folder:
                wt_path += folder + '/'
        if not wt_path.endswith("/UserSights/"):
            wt_path += "UserSights/"
    try:
        os.mkdir(wt_path)
        _output("Created folder " + wt_path, 1)
    except:
        pass
    _output("Writing in %s" % wt_path, 1)

    # Loading the table
    wrong_strings = []
    workbook = openpyxl.load_workbook("data.xlsx")

    generator.insert_str = dict[str, str]()
    for sheet_name in workbook.sheetnames:  # Iterating sheets
        if sheets and sheet_name not in sheets:  # Checking allowed list
            continue
        sheet = workbook[sheet_name]
        _output("\nReading " + sheet_name, 1)
        wrong_strings.append(0)
        empty_rows = 0
        row_num = 0

        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=9, values_only=True):  # Iterating rows inside certain sheet
            row_num += 1
            try:
                if row.count(None) == len(row):  # Checking emptiness
                    empty_rows += 1
                    continue
                elif empty_rows:
                    _output(str(empty_rows) + " empty rows", 0)
                    empty_rows = 0
                _output(str(row), 0)
                coords = list(map(lambda string: list(map(float, string.split(','))), row[5].split(';')))  # Reading first coords
                for i in 6, 7, 8:  # Continue reading coords
                    if row[i] is not None:
                        cur = list(map(lambda string: list(map(float, string.split(','))), row[i].split(';')))
                        for j in range(len(coords)):
                            for k in 0, 1:  # Coords summation
                                if i == 6:
                                    coords[j][k] += cur[j][k]
                                else:
                                    coords[j][k] -= cur[j][k]
                _output(str(coords), 0)
                # Removing 3% speed from AP, 5% from HE and HEAT
                type_list = row[4].split(';')
                speed_list = []
                for i in range(len(type_list)):
                    speed_list.append(int(str(row[2]).split(';')[i]))
                    if type_list[i] in ["sim_AP"]:
                        speed_list[-1] *= 0.97
                    if type_list[i] in ["sim_HEAT", "sim_HE"]:
                        speed_list[-1] *= 0.95
                # Create sight using generator
                _output(generator.generator(wt_path + row[0], speed_list, float(row[3]), type_list, coords, list(map(int, str(row[1]).split(';')))), 0)
            except:  # If something went wrong
                wrong_strings[-1] += 1
                _output("Wrong string format. Sheet: " + sheet_name + " Row: " + str(row_num), 1)

        _output(str(empty_rows) + " empty rows", 0)
        _output(str(wrong_strings[-1]) + " errors", 1)

    try:
        res = generator.save_presets()
    except:
        res = "\nError saving presets!\n"
    _output(res, 1)
    _output("Working directory was " + wt_path, 1)
    _output("Execution ended with " + str(sum(wrong_strings)) + " errors\n", 2)
    if MODE <= 2:
        _input("Press Enter to exit")


def cleaner(MODE, remove_all_tanks=False, _print=print, _input=input):
    def _output(string, severity=1):
        """
        Function that check should some text be moved to output or not.
        :param string: text that potentially should be moved to output
        :param severity: integer severity of this text
        """
        if MODE <= severity:
            _print(string)

    # Loading path from json and making it correct
    with open("settings.json", 'r') as f:
        wt_path = json.load(f)["path"].replace('\\', '/')
    if wt_path == "":
        wt_path = os.path.dirname(os.path.realpath("settings.json")) + "/UserSights/"
    else:
        wt_path_list = wt_path.split('/')
        wt_path = '/' if wt_path[0] == '/' else ""
        for folder in wt_path_list:
            if folder:
                wt_path += folder + '/'
        if not wt_path.endswith("/UserSights/"):
            wt_path += "UserSights/"

    _output("Deleting all from " + wt_path + '\n', 1)
    for dir_name in os.listdir(wt_path):
        if remove_all_tanks or dir_name != "all_tanks":
            try:
                _output("Deleting " + dir_name, 0)
                shutil.rmtree(os.path.join(wt_path, dir_name))
            except:
                pass
    _output("Deleted " + wt_path + '\n', 1)


if __name__ == "__main__":
    # Read all arguments from terminal and run main function
    parser = argparse.ArgumentParser(description="Creates UserSights folder with WarThunder sights.")
    parser.add_argument("-m", "--mode", help="Output mode. Development - 0 ; Normal - 1 (default) ; Silent - 2 ; Full silent - 3", default=1)
    MODE = int(vars(parser.parse_args())["mode"])

    reader(MODE, _print=print, _input=input)
